import openpyxl
from openpyxl.utils import get_column_letter
import html

# Nombre del archivo Excel de entrada (cámbialo según tu archivo)
input_file = "inut2.xlsx"

# Nombre del archivo HTML de salida
output_file = "inut.html"

def excel_to_html(input_file, output_file):
    try:
        # Cargar el archivo Excel
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active  # Usar la primera hoja (puedes cambiar a otra)

        # Obtener dimensiones de la hoja
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Obtener celdas combinadas
        merged_cells = sheet.merged_cells.ranges

        # Iniciar el contenido HTML con estilos
        html_content = [
            '<!DOCTYPE html>',
            '<html lang="es">',
            '<head>',
            '<meta charset="UTF-8">',
            '<title>Excel to HTML</title>',
            '<style>',
            'table { border-collapse: collapse; width: 100%; margin: 20px 0; }',
            'th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }',
            'th { background-color: #f2f2f2; font-weight: bold; }',
            '.merged { background-color: #f0f0f0; }',
            '</style>',
            '</head>',
            '<body>',
            '<table>'
        ]

        # Generar filas y celdas
        for row in range(1, max_row + 1):
            html_content.append('<tr>')
            col_index = 1
            while col_index <= max_col:
                cell = sheet.cell(row=row, column=col_index)
                value = cell.value if cell.value is not None else ""
                value = html.escape(str(value))  # Escapar caracteres especiales

                # Verificar si la celda está en un rango combinado
                is_merged = False
                rowspan = 1
                colspan = 1
                for merged_range in merged_cells:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        # Calcular rowspan y colspan
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        colspan = merged_range.max_col - merged_range.min_col + 1
                        break

                # Si la celda está combinada, solo incluir la primera celda del rango
                if is_merged:
                    # Verificar si esta es la celda inicial del rango combinado
                    for merged_range in merged_cells:
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            html_content.append(
                                f'<td class="merged" rowspan="{rowspan}" colspan="{colspan}">{value}</td>'
                            )
                            col_index += colspan
                            break
                    else:
                        col_index += 1
                else:
                    # Celda normal
                    tag = 'th' if row == 1 else 'td'  # Usar <th> para la primera fila
                    html_content.append(f'<{tag}>{value}</{tag}>')
                    col_index += 1
            html_content.append('</tr>')

        # Cerrar etiquetas HTML
        html_content.extend(['</table>', '</body>', '</html>'])

        # Guardar el archivo HTML
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html_content))

        print(f"Archivo HTML generado exitosamente: {output_file}")

    except FileNotFoundError:
        print(f"Error: El archivo {input_file} no se encontró.")
    except Exception as e:
        print(f"Error: {str(e)}")

# Ejecutar la función
excel_to_html(input_file, output_file)